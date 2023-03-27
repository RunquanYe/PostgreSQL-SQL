#!/usr/bin/env python
from openpyxl import Workbook
import psycopg2

'''
Programmer: Runquan Ye
'''

schemaName =  'ews'
size = '20'

tables = ['12_990', '12_990ez', '12_990pf', '13_990', '13_990pf', '13_EZ', '14_990', '14_990pf', '14_EZ', '15eofinextract990', '15eofinextract990pf', '15eofinextractEZ', '16eofinextract990', '16eofinextract990pf', '16eofinextractez', '17eofinextract990', '17eofinextractEZ', '18eoextractez', '18eoextract990', '19eoextract990', '19eoextractez']

column_names = []
colName = 'EIN'

existTable = []
dontExistTable = []

try:
    book = Workbook()

    # get the reference to the active sheet
    sheet = book.active
    sheet.name = 'Column Sample'

    connection = psycopg2.connect(
        host = "localhost",
        database = "cri",
        port = "5432")


    with connection:
        with connection.cursor() as cursor:
            rowIndex = 1
            for tName in tables:
                checkComand = f"SELECT EXISTS (SELECT 1 FROM information_schema.columns WHERE table_schema='{schemaName}' AND table_name='{tName}' AND column_name='{colName}');"
                cursor.execute(checkComand)
        
                checkExist = bool(*(row[0] for row in cursor))
                
                sheet.cell(row=1, column=1).value = "Table Col Result"
                sheet.cell(row=2, column=1).value = "Table Name"
                sheet.cell(row=2, column=2).value = "Column Name"
                sheet.cell(row=2, column=3).value = "Min Result"
                sheet.cell(row=2, column=4).value = "Min Length"
                sheet.cell(row=2, column=5).value = "Max Result"
                sheet.cell(row=2, column=6).value = "Max Length"
                if checkExist:
                    command = 'select "{0}" from {1}."{2}" limit {3};'.format(colName, schemaName, tName, size)
                    cursor.execute(command)
                    result = [row[0] for row in cursor]
                    sheet.cell(row=rowIndex + 2, column=1).value = tName
                    sheet.cell(row=rowIndex + 2, column=2).value = colName
                    sheet.cell(row=rowIndex + 2, column=3).value = min(result, key=len)
                    sheet.cell(row=rowIndex + 2, column=4).value = len(min(result, key=len))
                    sheet.cell(row=rowIndex + 2, column=5).value = max(result, key=len)
                    sheet.cell(row=rowIndex + 2, column=6).value = len(max(result, key=len))
                    sheet.cell(row=rowIndex + 2, column=7).value = ", ".join(str(c) for c in result)
                    existTable.append(tName)
                    rowIndex += 1
                else:
                    # print(f"Table {tName} doesn't have {colName} Column.")
                    pass

except (Exception, psycopg2.Error) as error :
    print ("Error: ", error)

finally:
    dontExistTable = [t for t in tables if t not in existTable]
    sheet.cell(row=rowIndex+3, column=1).value = 'Result:'
    sheet.cell(row=rowIndex+4, column=1).value = f'Total Table: {len(tables)}, Exist Col Table: {len(existTable)}, Miss Col Table: {len(dontExistTable)}'
    sheet.cell(row=rowIndex+5, column=1).value = f"Exist Col Table: " + ", ".join(str(c) for c in existTable)
    sheet.cell(row=rowIndex+6, column=1).value = f"Miss Col Table: " + ", ".join(str(c) for c in dontExistTable)
    
    book.save("TableColSample.xlsx")

    #closing database connection.
    if(connection):
        cursor.close()
        connection.close()
        print("PostgreSQL connection is closed")
